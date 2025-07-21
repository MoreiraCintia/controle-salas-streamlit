import pandas as pd
import streamlit as st
import io
from openpyxl import load_workbook
from datetime import datetime, timedelta

# 📁 Caminho do seu arquivo
ARQUIVO_EXCEL = "Planejamento e Custos - 2025.xlsm"

# 🎯 Função para extrair os dados da aba Planejamento e montar o controle de salas
@st.cache_data
def gerar_controle_de_salas():
    wb = load_workbook(filename=ARQUIVO_EXCEL, data_only=True)
    ws = wb["Planejamento"]

    dados = []
    dias_semana_col = {0: 'I', 1: 'J', 2: 'K', 3: 'L', 4: 'M'}  # Segunda a sexta

    for row in range(2, ws.max_row + 1):
        curso = ws[f"B{row}"].value
        periodo = str(ws[f"D{row}"].value).strip().capitalize()
        segunda = ws[f"I{row}"].value
        terca = ws[f"J{row}"].value
        quarta = ws[f"K{row}"].value
        quinta = ws[f"L{row}"].value
        sexta = ws[f"M{row}"].value
        dias_x = [segunda, terca, quarta, quinta, sexta]
        inicio = ws[f"S{row}"].value
        fim = ws[f"T{row}"].value
        sala = str(ws[f"AF{row}"].value).strip().upper()

        if not (curso and inicio and fim and sala and periodo):
            continue

        data_atual = inicio
        while data_atual <= fim:
            dia_semana_idx = data_atual.weekday()  # 0 = segunda
            if dia_semana_idx < 5 and str(dias_x[dia_semana_idx]).strip().lower() == 'x':
                dados.append({
                    "Data": data_atual,
                    "Sala": sala,
                    "Período": periodo,
                    "Curso": curso,
                    "Status": "Ocupado"
                })
            data_atual += timedelta(days=1)

    # Obter todas as combinações possíveis de datas, salas e períodos
    datas = sorted(set(d["Data"] for d in dados))
    salas = sorted(set(d["Sala"] for d in dados))
    periodos = ["Manhã", "Tarde", "Noite"]
    combinacoes = []

    for data in datas:
        for sala in salas:
            for periodo in periodos:
                # Verifica se essa combinação já está ocupada
                encontrado = next((d for d in dados if d["Data"] == data and d["Sala"] == sala and d["Período"] == periodo), None)
                if encontrado:
                    combinacoes.append(encontrado)
                else:
                    combinacoes.append({
                        "Data": data,
                        "Sala": sala,
                        "Período": periodo,
                        "Curso": "",
                        "Status": "Livre"
                    })

    return pd.DataFrame(combinacoes)

# 🔁 Carrega os dados
df = gerar_controle_de_salas()

# 🎨 Interface do Streamlit
st.title("📊 Controle de Ocupação das Salas")

# 📍 Filtros independentes
col1, col2 = st.columns(2)

# Obtemos a menor e maior data do DataFrame
data_min = df["Data"].min().date()
data_max = df["Data"].max().date()

with col1:
    datas_filtro = st.date_input("📅 Período", value=(data_min, data_max), min_value=data_min, max_value=data_max)

with col2:
    status_filtro = st.selectbox("📌 Status", options=["Todos", "Livre", "Ocupado"])


col3, col4 = st.columns(2)
with col3:
    sala_filtro = st.selectbox("🏫 Sala", options=["Todas"] + sorted(df["Sala"].unique().tolist()))
with col4:
    periodo_filtro = st.selectbox("⏰ Período", options=["Todos"] + ["Manhã", "Tarde", "Noite"])

# 🎯 Aplica filtros
filtro = df.copy()
# Aplica filtro de período (se ambas datas forem escolhidas)
if datas_filtro and all(datas_filtro):
    inicio, fim = datas_filtro
    filtro = filtro[(filtro["Data"] >= pd.to_datetime(inicio)) & (filtro["Data"] <= pd.to_datetime(fim))]
if status_filtro != "Todos":
    filtro = filtro[filtro["Status"] == status_filtro]
if sala_filtro != "Todas":
    filtro = filtro[filtro["Sala"] == sala_filtro]
if periodo_filtro != "Todos":
    filtro = filtro[filtro["Período"] == periodo_filtro]

filtro["Data"] = pd.to_datetime(filtro["Data"], errors="coerce").dt.strftime("%d/%m/%Y")


# 🧾 Resultado
filtro["Data"] = pd.to_datetime(filtro["Data"], errors="coerce").dt.strftime("%d/%m/%Y")
# Cria uma cópia apenas para exibição, com a coluna "Data" formatada
df_exibicao = filtro.copy()
df_exibicao["Data"] = pd.to_datetime(df_exibicao["Data"]).dt.strftime("%d/%m/%Y")

st.dataframe(df_exibicao.sort_values(["Data", "Sala", "Período"]))


# Gera um arquivo Excel em memória
output = io.BytesIO()
with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
    filtro.to_excel(writer, index=False, sheet_name="Controle de Salas")
output.seek(0)

import io
st.set_page_config(layout="wide"

# Gera um arquivo Excel em memória
output = io.BytesIO()
with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
    filtro.to_excel(writer, index=False, sheet_name="Controle de Salas")
output.seek(0)

# Botão para download do Excel
st.download_button(
    label="📥 Baixar como Excel",
    data=output,
    file_name="controle_salas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
